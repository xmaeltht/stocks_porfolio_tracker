#!/usr/bin/env python3
"""
Portfolio Monitor
-----------------
Reads holdings from portfolio_config.json, fetches live market data via yfinance,
and generates a professional Excel report (portfolio_report.xlsx).

Run this script hourly during market hours or trigger it from the dashboard.
"""

import json
import os
import sys
from datetime import datetime

try:
    import yfinance as yf
except ImportError:
    print("Installing yfinance...")
    os.system(f"{sys.executable} -m pip install yfinance --break-system-packages -q")
    import yfinance as yf

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl --break-system-packages -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "portfolio_config.json")
OUTPUT_FILE = os.path.join(BASE_DIR, "portfolio_report.xlsx")


# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
def load_config():
    if not os.path.exists(CONFIG_FILE):
        default = {"name": "My Portfolio", "alert_threshold": 5.0, "holdings": []}
        with open(CONFIG_FILE, "w") as f:
            json.dump(default, f, indent=2)
        return default
    with open(CONFIG_FILE) as f:
        return json.load(f)


# ─────────────────────────────────────────────
# DATA FETCHING
# ─────────────────────────────────────────────
def fetch_data(holdings, alert_threshold=5.0):
    results = []
    for h in holdings:
        ticker = h.get("ticker", "").strip().upper()
        shares = float(h.get("shares", 0))
        avg_cost = float(h.get("avg_cost", 0))
        if not ticker:
            continue
        try:
            stock = yf.Ticker(ticker)
            info = stock.info
            price = (
                info.get("regularMarketPrice")
                or info.get("currentPrice")
                or info.get("previousClose")
                or 0
            )
            price = float(price or 0)
            change = float(info.get("regularMarketChange") or 0)
            change_pct = float(info.get("regularMarketChangePercent") or 0)
            value = price * shares
            cost_basis = avg_cost * shares
            pnl = (value - cost_basis) if avg_cost else 0
            pnl_pct = (pnl / cost_basis * 100) if cost_basis else 0

            results.append({
                "ticker": ticker,
                "name": info.get("shortName", ticker),
                "shares": shares,
                "avg_cost": avg_cost,
                "price": price,
                "change": change,
                "change_pct": change_pct,
                "value": value,
                "pnl": pnl,
                "pnl_pct": pnl_pct,
                "volume": float(info.get("regularMarketVolume") or 0),
                "market_cap": float(info.get("marketCap") or 0),
                "week52_high": float(info.get("fiftyTwoWeekHigh") or 0),
                "week52_low": float(info.get("fiftyTwoWeekLow") or 0),
                "pe_ratio": float(info.get("trailingPE") or 0),
                "alert": abs(change_pct) >= alert_threshold,
            })
            status = "🚨" if abs(change_pct) >= alert_threshold else "✅"
            arrow = "↑" if change >= 0 else "↓"
            print(f"  {status} {ticker:6s}  ${price:>9.2f}  {arrow} {change_pct:+.2f}%")
        except Exception as e:
            print(f"  ⚠️  {ticker}: Error - {e}")
            results.append({
                "ticker": ticker, "name": ticker,
                "shares": shares, "avg_cost": avg_cost,
                "price": 0, "change": 0, "change_pct": 0,
                "value": 0, "pnl": 0, "pnl_pct": 0,
                "volume": 0, "market_cap": 0,
                "week52_high": 0, "week52_low": 0, "pe_ratio": 0,
                "alert": False,
            })
    return results


# ─────────────────────────────────────────────
# EXCEL GENERATION
# ─────────────────────────────────────────────
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def hdr_cell(ws, row, col, value, bg="1F3864", fg="FFFFFF", bold=True, size=10, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    return c


def data_cell(ws, row, col, value, num_fmt=None, color=None, bg=None, align="right"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, color=color or "000000")
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = thin_border()
    if num_fmt:
        c.number_format = num_fmt
    return c


def generate_excel(config, data, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_value = sum(h["value"] for h in data)
    total_cost = sum(h["avg_cost"] * h["shares"] for h in data)
    total_pnl = sum(h["pnl"] for h in data)
    total_pnl_pct = (total_pnl / total_cost * 100) if total_cost else 0
    day_pnl = sum(h["change"] * h["shares"] for h in data)
    alerts = [h for h in data if h["alert"]]
    best = max(data, key=lambda x: x["change_pct"], default=None)
    worst = min(data, key=lambda x: x["change_pct"], default=None)
    threshold = config.get("alert_threshold", 5.0)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── SUMMARY SHEET ────────────────────────────────
    ws = wb.create_sheet("📊 Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"  {config.get('name', 'My Portfolio')}  ·  Market Monitor"
    t.font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", start_color="0D1117")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 42

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value = f"  Last updated: {generated}   |   Holdings: {len(data)}   |   Alert threshold: ±{threshold}%"
    sub.font = Font(name="Arial", size=9, italic=True, color="888888")
    sub.fill = PatternFill("solid", start_color="0D1117")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 10

    # KPI cards (row 4–5)
    kpis = [
        ("PORTFOLIO VALUE", f"${total_value:,.2f}", "1F3864", "BDD7EE"),
        ("TODAY'S P&L", f"{'+'if day_pnl>=0 else ''}${day_pnl:,.2f}", "375623" if day_pnl >= 0 else "7B2D00", "E2EFDA" if day_pnl >= 0 else "FCE4EC"),
        ("TOTAL RETURN", f"{'+'if total_pnl_pct>=0 else ''}{total_pnl_pct:.2f}%", "375623" if total_pnl_pct >= 0 else "7B2D00", "E2EFDA" if total_pnl_pct >= 0 else "FCE4EC"),
        ("ACTIVE ALERTS", f"{len(alerts)} stock{'s' if len(alerts)!=1 else ''}", "7B2D00" if alerts else "1F3864", "FCE4EC" if alerts else "BDD7EE"),
    ]
    kpi_cols = [1, 2, 4, 5]
    for (label, value, fg, bg), col in zip(kpis, [1, 2, 4, 5]):
        for r in [4, 5]:
            ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=bg)
            ws.cell(row=r, column=col).border = thin_border()
        ws.cell(row=4, column=col).value = label
        ws.cell(row=4, column=col).font = Font(name="Arial", bold=True, size=9, color=fg)
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=col).value = value
        ws.cell(row=5, column=col).font = Font(name="Arial", bold=True, size=14, color=fg)
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 28

    ws.row_dimensions[6].height = 10

    # Best / Worst
    row = 7
    hdr_cell(ws, row, 1, "BEST PERFORMER", "375623", align="left")
    hdr_cell(ws, row, 2, "WORST PERFORMER", "7B2D00", align="left")
    row += 1
    if best:
        ws.cell(row=row, column=1).value = f"{best['ticker']}  ↑ +{best['change_pct']:.2f}%"
        ws.cell(row=row, column=1).font = Font(name="Arial", bold=True, size=11, color="375623")
    if worst:
        ws.cell(row=row, column=2).value = f"{worst['ticker']}  ↓ {worst['change_pct']:.2f}%"
        ws.cell(row=row, column=2).font = Font(name="Arial", bold=True, size=11, color="C00000")

    ws.row_dimensions[8].height = 22
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28

    # ── HOLDINGS SHEET ───────────────────────────────
    ws2 = wb.create_sheet("📋 Holdings")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:P1")
    th = ws2["A1"]
    th.value = f"  Holdings Detail — {generated}"
    th.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    th.fill = PatternFill("solid", start_color="1F3864")
    th.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 32

    col_defs = [
        ("Ticker",          8,  "@",              "left"),
        ("Company",         26, "@",              "left"),
        ("Shares",          9,  "#,##0.000",      "right"),
        ("Avg Cost",        11, "$#,##0.00",      "right"),
        ("Price",           11, "$#,##0.00",      "right"),
        ("Day Chg $",       11, '$#,##0.00;[Red]($#,##0.00)',  "right"),
        ("Day Chg %",       10, '0.00%;[Red]-0.00%',           "right"),
        ("Mkt Value",       13, "$#,##0.00",      "right"),
        ("Total P&L",       12, '$#,##0.00;[Red]($#,##0.00)',  "right"),
        ("P&L %",           9,  '0.00%;[Red]-0.00%',           "right"),
        ("52W High",        10, "$#,##0.00",      "right"),
        ("52W Low",         10, "$#,##0.00",      "right"),
        ("P/E",             8,  "0.0",            "right"),
        ("Volume",          13, "#,##0",          "right"),
        ("Mkt Cap",         12, '$#,##0,,,"B"',   "right"),
        ("Alert",           8,  "@",              "center"),
    ]

    for col_idx, (label, width, fmt, align) in enumerate(col_defs, 1):
        hdr_cell(ws2, 2, col_idx, label, bg="1F3864")
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[2].height = 20
    ws2.freeze_panes = "A3"

    for ri, h in enumerate(data, 3):
        row_bg = "F0F4F8" if ri % 2 == 0 else "FFFFFF"
        alert_bg = "FFF3CD" if h["alert"] else None
        bg = alert_bg or row_bg

        chg_color = "375623" if h["change"] >= 0 else "C00000"
        pnl_color = "375623" if h["pnl"] >= 0 else "C00000"

        row_vals = [
            (h["ticker"],      "@",              "000000", "left"),
            (h["name"],        "@",              "000000", "left"),
            (h["shares"],      "#,##0.000",      "000000", "right"),
            (h["avg_cost"],    "$#,##0.00",      "000000", "right"),
            (h["price"],       "$#,##0.00",      "000000", "right"),
            (h["change"],      '$#,##0.00;[Red]($#,##0.00)', chg_color, "right"),
            (h["change_pct"]/100, '0.00%;[Red]-0.00%', chg_color, "right"),
            (h["value"],       "$#,##0.00",      "000000", "right"),
            (h["pnl"],         '$#,##0.00;[Red]($#,##0.00)', pnl_color, "right"),
            (h["pnl_pct"]/100, '0.00%;[Red]-0.00%', pnl_color, "right"),
            (h["week52_high"], "$#,##0.00",      "000000", "right"),
            (h["week52_low"],  "$#,##0.00",      "000000", "right"),
            (h["pe_ratio"],    "0.0",            "000000", "right"),
            (h["volume"],      "#,##0",          "000000", "right"),
            (h["market_cap"],  '$#,##0,,,"B"',   "000000", "right"),
            ("🚨 ALERT" if h["alert"] else "—", "@", "C00000" if h["alert"] else "888888", "center"),
        ]
        for ci, (val, fmt, color, align) in enumerate(row_vals, 1):
            c = ws2.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10, color=color)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = thin_border()
            c.number_format = fmt
        ws2.row_dimensions[ri].height = 18

    # ── ALERTS SHEET ────────────────────────────────
    ws3 = wb.create_sheet("🚨 Alerts")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:G1")
    ta = ws3["A1"]
    ta.value = f"  Market Alerts — Stocks Moving ±{threshold}%   ({generated})"
    ta.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ta.fill = PatternFill("solid", start_color="C00000" if alerts else "375623")
    ta.alignment = Alignment(horizontal="left", vertical="center")
    ws3.row_dimensions[1].height = 32

    if not alerts:
        ws3["A2"].value = f"✅  No stocks have moved more than ±{threshold}% today."
        ws3["A2"].font = Font(name="Arial", size=11, color="375623")
        ws3["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws3.row_dimensions[2].height = 28
    else:
        alert_cols = [
            ("Ticker", 8, "@", "left"),
            ("Company", 26, "@", "left"),
            ("Price", 12, "$#,##0.00", "right"),
            ("Day Chg $", 12, '$#,##0.00;[Red]($#,##0.00)', "right"),
            ("Day Chg %", 12, '0.00%;[Red]-0.00%', "right"),
            ("Shares", 9, "#,##0.000", "right"),
            ("Day Impact", 13, '$#,##0.00;[Red]($#,##0.00)', "right"),
        ]
        for ci, (lbl, w, fmt, align) in enumerate(alert_cols, 1):
            hdr_cell(ws3, 2, ci, lbl, bg="C00000")
            ws3.column_dimensions[get_column_letter(ci)].width = w
        ws3.row_dimensions[2].height = 20

        for ri, h in enumerate(alerts, 3):
            bg = "E2EFDA" if h["change"] >= 0 else "FCE4EC"
            fg = "375623" if h["change"] >= 0 else "C00000"
            impact = h["change"] * h["shares"]
            alert_row = [
                (h["ticker"],          "@",    fg, "left"),
                (h["name"],            "@",    "000000", "left"),
                (h["price"],           "$#,##0.00", "000000", "right"),
                (h["change"],          '$#,##0.00;[Red]($#,##0.00)', fg, "right"),
                (h["change_pct"]/100,  '0.00%;[Red]-0.00%', fg, "right"),
                (h["shares"],          "#,##0.000", "000000", "right"),
                (impact,               '$#,##0.00;[Red]($#,##0.00)', fg, "right"),
            ]
            for ci, (val, fmt, color, align) in enumerate(alert_row, 1):
                c = ws3.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", bold=True, size=10, color=color)
                c.fill = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal=align, vertical="center")
                c.border = thin_border()
                c.number_format = fmt
            ws3.row_dimensions[ri].height = 20

    wb.save(output_path)
    print(f"\n✅  Excel report saved → {output_path}")


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'═'*52}")
    print(f"  📊 Portfolio Monitor  ·  {now}")
    print(f"{'═'*52}")

    config = load_config()
    holdings = config.get("holdings", [])

    if not holdings:
        print("\n⚠️  No holdings found in portfolio_config.json")
        print("   Open portfolio_dashboard.html to add your stocks,")
        print("   then export the config and save it as portfolio_config.json\n")
        return

    threshold = config.get("alert_threshold", 5.0)
    print(f"\n📡 Fetching {len(holdings)} holdings  (alert threshold: ±{threshold}%)\n")
    data = fetch_data(holdings, threshold)

    total_value = sum(h["value"] for h in data)
    day_pnl = sum(h["change"] * h["shares"] for h in data)
    alerts = [h for h in data if h["alert"]]

    print(f"\n{'─'*52}")
    print(f"  💼  Portfolio Value :  ${total_value:>12,.2f}")
    print(f"  📈  Today's P&L    :  ${day_pnl:>+12,.2f}")
    if alerts:
        print(f"\n  🚨  {len(alerts)} alert(s) — stocks moved ±{threshold}%:")
        for h in alerts:
            print(f"       {h['ticker']:6s}  {h['change_pct']:+.2f}%")
    print(f"{'─'*52}\n")

    generate_excel(config, data, OUTPUT_FILE)


if __name__ == "__main__":
    main()
